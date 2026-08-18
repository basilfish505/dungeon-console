"""Load and write the item types spreadsheet.

Fill item_types.xlsx (Items sheet, one row per item type), save, and
restart the server. Extra columns are ignored so the sheet can grow later.
"""

from __future__ import annotations

from pathlib import Path

from item_types.base import ItemTypeDef
from item_types.registry import register_item_type

PROJECT_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_XLSX_PATH = PROJECT_ROOT / 'item_types.xlsx'

COLUMNS = [
    'item_id',
    'name',
    'description',
    'price_pqg',
    'image',
]

HEADER_COMMENTS = {
    'item_id': 'Stable machine id, lowercase snake_case (e.g. healing_potion). Required.',
    'name': 'Display name shown to players. Required.',
    'description': 'Inspect / inventory flavor text.',
    'price_pqg': 'Price in PermaQuest Gold (integer). Default 0.',
    'image': 'Sprite path or filename. Blank = /static/items/sprites/{item_id}.png',
}

TEST_ITEMS = [
    {
        'item_id': 'healing_potion',
        'name': 'Healing Potion',
        'description': 'Restores a small amount of health.',
        'price_pqg': 25,
        'image': 'healing_potion.png',
    },
    {
        'item_id': 'torch',
        'name': 'Torch',
        'description': 'A wooden torch. Useful in dark places.',
        'price_pqg': 5,
        'image': 'torch.png',
    },
    {
        'item_id': 'bread',
        'name': 'Bread',
        'description': 'A loaf of dense travel bread.',
        'price_pqg': 3,
        'image': 'bread.png',
    },
    {
        'item_id': 'rope',
        'name': 'Rope',
        'description': 'Fifty feet of sturdy hemp rope.',
        'price_pqg': 15,
        'image': 'rope.png',
    },
    {
        'item_id': 'antidote',
        'name': 'Antidote',
        'description': 'Cures common poisons.',
        'price_pqg': 20,
        'image': 'antidote.png',
    },
]


def _blank(value):
    if value is None:
        return True
    if isinstance(value, str) and not value.strip():
        return True
    return False


def _int(value, default):
    if _blank(value):
        return default
    return int(float(value))


def _str_or_none(value):
    if _blank(value):
        return None
    return str(value).strip()


def _normalize_header(value):
    if value is None:
        return ''
    return str(value).strip()


def row_to_typedef(row):
    """Build an ItemTypeDef from a column->value mapping. None if no item_id.

    Extra spreadsheet columns are ignored (row.get only known fields).
    """
    item_id = _str_or_none(row.get('item_id'))
    if not item_id:
        return None
    return ItemTypeDef(
        item_id=item_id,
        name=_str_or_none(row.get('name')) or item_id,
        description=_str_or_none(row.get('description')),
        price_pqg=_int(row.get('price_pqg'), 0),
        image=_str_or_none(row.get('image')),
    )


def iter_xlsx_rows(path):
    from openpyxl import load_workbook

    wb = load_workbook(Path(path), data_only=True, read_only=True)
    try:
        ws = wb['Items'] if 'Items' in wb.sheetnames else wb.active
        rows = ws.iter_rows(values_only=True)
        header_row = next(rows, None)
        if not header_row:
            return
        headers = [_normalize_header(h) for h in header_row]
        for values in rows:
            yield {
                headers[i]: values[i] if i < len(values) else None
                for i in range(len(headers))
                if headers[i]
            }
    finally:
        wb.close()


def typedefs_from_rows(rows):
    types = []
    for row in rows:
        type_def = row_to_typedef(row)
        if type_def is not None:
            types.append(type_def)
    return types


def load_item_sheet(path, register=True):
    """Load item types from a .xlsx workbook. Returns the ItemTypeDef list."""
    path = Path(path)
    if not path.is_file():
        raise FileNotFoundError(path)
    if path.suffix.lower() != '.xlsx':
        raise ValueError(f'Item sheet must be .xlsx, got: {path.suffix}')
    types = typedefs_from_rows(iter_xlsx_rows(path))
    if register:
        for type_def in types:
            register_item_type(type_def)
    return types


def load_default_item_sheet():
    """Load project-root item_types.xlsx. Returns [] if missing or unloadable."""
    if not DEFAULT_XLSX_PATH.is_file():
        print(
            f'[item_types] missing {DEFAULT_XLSX_PATH}; '
            'no spreadsheet item definitions loaded'
        )
        return []
    try:
        return load_item_sheet(DEFAULT_XLSX_PATH)
    except ImportError as exc:
        print(
            f'[item_types] cannot load {DEFAULT_XLSX_PATH}: {exc}. '
            'Install openpyxl in this Python environment '
            '(pip install openpyxl)'
        )
        return []
    except Exception as exc:
        print(f'[item_types] failed to load {DEFAULT_XLSX_PATH}: {exc}')
        return []


def write_item_xlsx(path=None, extra_rows=None):
    """Create the fill-in item workbook (Items + Instructions + Field reference)."""
    from openpyxl import Workbook
    from openpyxl.comments import Comment
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    path = Path(path) if path else DEFAULT_XLSX_PATH
    wb = Workbook()

    thin = Border(
        left=Side(style='thin', color='C4B8A8'),
        right=Side(style='thin', color='C4B8A8'),
        top=Side(style='thin', color='C4B8A8'),
        bottom=Side(style='thin', color='C4B8A8'),
    )
    header_fill = PatternFill('solid', fgColor='3D2B1F')
    header_font = Font(bold=True, color='F4E8D0', name='Calibri', size=11)
    example_fill = PatternFill('solid', fgColor='E6F4EA')
    wrap = Alignment(wrap_text=True, vertical='center')

    ws = wb.active
    ws.title = 'Items'
    for col_idx, header in enumerate(COLUMNS, 1):
        cell = ws.cell(1, col_idx, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin
        note = HEADER_COMMENTS.get(header)
        if note:
            cell.comment = Comment(note, 'dungeon-console')

    data_rows = list(TEST_ITEMS)
    if extra_rows:
        data_rows.extend(extra_rows)

    for row_idx, row in enumerate(data_rows, 2):
        for col_idx, header in enumerate(COLUMNS, 1):
            cell = ws.cell(row_idx, col_idx, row.get(header, ''))
            cell.border = thin
            cell.alignment = wrap
            if row_idx <= 6:
                cell.fill = example_fill

    blank_count = max(0, 40 - len(data_rows))
    last_data = 1 + len(data_rows) + blank_count
    for row_idx in range(2 + len(data_rows), last_data + 1):
        for col_idx in range(1, len(COLUMNS) + 1):
            cell = ws.cell(row_idx, col_idx, None)
            cell.border = thin

    widths = {
        'item_id': 18,
        'name': 18,
        'description': 48,
        'price_pqg': 12,
        'image': 28,
    }
    for col_idx, header in enumerate(COLUMNS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(header, 12)

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(COLUMNS))}{last_data}'
    ws.row_dimensions[1].height = 22

    ins = wb.create_sheet('Instructions')
    ins['A1'] = 'How to use'
    ins['A1'].font = Font(bold=True, size=14, color='3D2B1F')
    instructions = [
        '',
        '1. Fill one row per item type on the Items sheet.',
        '2. Required: item_id. Also fill name, description, and price_pqg.',
        '3. Leave image blank to use /static/items/sprites/{item_id}.png',
        '4. Extra columns may be added later; the loader ignores unknown headers.',
        '5. Save this workbook, then restart the game to reload definitions.',
        '',
        'Art files',
        'Put PNG files at static/items/sprites/{item_id}.png',
        'or set image to a custom path under /static/.',
    ]
    for idx, line in enumerate(instructions, 2):
        ins[f'A{idx}'] = line
        if line in ('Art files',):
            ins[f'A{idx}'].font = Font(bold=True, size=12, color='3D2B1F')
    ins.column_dimensions['A'].width = 90

    ref = wb.create_sheet('Field reference')
    ref_headers = ('column', 'required', 'type', 'notes')
    for col_idx, header in enumerate(ref_headers, 1):
        cell = ref.cell(1, col_idx, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin
    fields = [
        ('item_id', 'yes', 'string', 'Unique machine id'),
        ('name', 'yes', 'string', 'Display name'),
        ('description', 'no', 'string', 'Inventory text'),
        ('price_pqg', 'no', 'int', 'Price in PQG; default 0'),
        ('image', 'no', 'path/filename', 'Blank = default sprite path'),
    ]
    for row_idx, values in enumerate(fields, 2):
        for col_idx, value in enumerate(values, 1):
            cell = ref.cell(row_idx, col_idx, value)
            cell.border = thin
            cell.alignment = wrap
    ref.column_dimensions['A'].width = 16
    ref.column_dimensions['B'].width = 12
    ref.column_dimensions['C'].width = 16
    ref.column_dimensions['D'].width = 40

    wb.save(path)
    return path
