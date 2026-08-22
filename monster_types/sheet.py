"""Load and write the monster species spreadsheet.

Fill monster_types.xlsx (Monsters sheet, one row per species), save, and
restart the server. Rows with a type_id are registered and can spawn.
"""

from __future__ import annotations

from pathlib import Path

from character_stats import ATTRIBUTE_KEYS
from monster_types.base import MonsterTypeDef
from monster_types.leveling import DEFAULT_LEVEL_SCALING, DEFAULT_MAX_LEVEL
from monster_types.registry import register_monster_type

PROJECT_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_XLSX_PATH = PROJECT_ROOT / 'monster_types.xlsx'

COLUMNS = [
    'type_id',
    'name',
    'description',
    'base_level',
    'max_level',
    'level_scaling',
    'str',
    'int',
    'wis',
    'chr',
    'dex',
    'agi',
    'acc',
    'base_mhp',
    'armour',
    'aggression',
    'speed',
    'activeness',
    'sight_range',
    'ability_ids',
    'sprite',
    'portrait',
    'spawn_weight',
    'spawn_notes',
]

HEADER_COMMENTS = {
    'type_id': 'Machine id, lowercase, unique (e.g. goblin). Required.',
    'name': 'Display name shown to players. Required.',
    'description': 'Inspect / flavor text.',
    'base_level': 'Default level when construction omits level= (tests/NPCs). Default 1.',
    'max_level': f'Highest random spawn level (inclusive). Default {DEFAULT_MAX_LEVEL}.',
    'level_scaling': (
        f'Bonus attribute points per level above 1. '
        f'Default {DEFAULT_LEVEL_SCALING}.'
    ),
    'str': 'Strength. Integer.',
    'int': 'Intelligence. Integer.',
    'wis': 'Wisdom. Integer.',
    'chr': 'Charisma. Integer.',
    'dex': 'Dexterity. Integer.',
    'agi': 'Agility. Integer.',
    'acc': 'ACC. Integer. Attacker hit chance vs defender Dexterity.',
    'base_mhp': 'Max HP at base level. Required.',
    'armour': 'Damage divisor in combat. 1 = full damage, 2 = half, etc. Default 1.',
    'aggression': '0-10. Low flees, ~5 wanders, high chases. Default 0.',
    'speed': '0-10. 0 never acts; any value above 0 gets one step per world round.',
    'activeness': '0-10. Chance to stay still when idle/neutral = 1 - activeness/10.',
    'sight_range': 'Chebyshev vision distance in tiles. Default 20.',
    'ability_ids': 'Comma-separated ability ids, or blank. Combat hooks not implemented yet.',
    'sprite': 'Leave blank for /static/monsters/sprites/{type_id}.png',
    'portrait': 'Leave blank for /static/monsters/portraits/{type_id}.png',
    'spawn_weight': 'Relative chance to appear when a monster spawns. 0 = never random-spawn.',
    'spawn_notes': 'Design notes only; not loaded into the game.',
}

TROLL_EXAMPLE = {
    'type_id': 'troll',
    'name': 'Troll',
    'description': 'A large, brutish creature that relies on strength and durability.',
    'base_level': 1,
    'max_level': DEFAULT_MAX_LEVEL,
    'level_scaling': DEFAULT_LEVEL_SCALING,
    'str': 8,
    'int': 3,
    'wis': 3,
    'chr': 2,
    'dex': 4,
    'agi': 4,
    'acc': 4,
    'base_mhp': 16,
    'armour': 1,
    'aggression': 0,
    'speed': 10,
    'activeness': 5,
    'sight_range': 20,
    'ability_ids': '',
    'sprite': '/static/monsters/sprites/troll.png',
    'portrait': '/static/monsters/portraits/troll.png',
    'spawn_weight': 1,
    'spawn_notes': 'Current default spawn. Add more rows below.',
}


def _blank(value):
    if value is None:
        return True
    if isinstance(value, str) and not value.strip():
        return True
    return False


def parse_ability_ids(value):
    if _blank(value):
        return []
    return [part.strip() for part in str(value).split(',') if part.strip()]


def _int(value, default):
    if _blank(value):
        return default
    return int(float(value))


def _float(value, default):
    if _blank(value):
        return default
    return float(value)


def _str_or_none(value):
    if _blank(value):
        return None
    return str(value).strip()


def row_to_typedef(row):
    """Build a MonsterTypeDef from a column->value mapping. None if no type_id."""
    type_id = _str_or_none(row.get('type_id'))
    if not type_id:
        return None
    name = _str_or_none(row.get('name')) or type_id
    attrs = {key: _int(row.get(key), 1) for key in ATTRIBUTE_KEYS}
    if not _blank(row.get('accuracy')) and _blank(row.get('acc')):
        attrs['acc'] = _int(row.get('accuracy'), 1)
    return MonsterTypeDef(
        type_id=type_id,
        name=name,
        description=_str_or_none(row.get('description')),
        base_level=_int(row.get('base_level'), 1),
        max_level=_int(row.get('max_level'), DEFAULT_MAX_LEVEL),
        level_scaling=_int(row.get('level_scaling'), DEFAULT_LEVEL_SCALING),
        base_attributes=attrs,
        base_mhp=_int(row.get('base_mhp'), 10),
        armour=_int(row.get('armour'), 1),
        aggression=_float(row.get('aggression'), 0),
        speed=_float(row.get('speed'), 10),
        activeness=_float(row.get('activeness'), 5),
        sight_range=_int(row.get('sight_range'), 20),
        ability_ids=parse_ability_ids(row.get('ability_ids')),
        sprite=_str_or_none(row.get('sprite')),
        portrait=_str_or_none(row.get('portrait')),
        spawn_weight=_float(row.get('spawn_weight'), 1),
    )


def _normalize_header(value):
    if value is None:
        return ''
    return str(value).strip()


def iter_xlsx_rows(path):
    from openpyxl import load_workbook

    wb = load_workbook(Path(path), data_only=True, read_only=True)
    try:
        ws = wb['Monsters'] if 'Monsters' in wb.sheetnames else wb.active
        rows = ws.iter_rows(values_only=True)
        header_row = next(rows, None)
        if not header_row:
            return
        headers = [_normalize_header(h) for h in header_row]
        for values in rows:
            yield {
                headers[i]: values[i] if i < len(values) else None
                for i in range(len(headers))
                if headers[i]
            }
    finally:
        wb.close()


def typedefs_from_rows(rows):
    types = []
    for row in rows:
        type_def = row_to_typedef(row)
        if type_def is not None:
            types.append(type_def)
    return types


def load_monster_sheet(path, register=True):
    """Load species from a .xlsx workbook. Returns the MonsterTypeDef list."""
    path = Path(path)
    if not path.is_file():
        raise FileNotFoundError(path)
    if path.suffix.lower() != '.xlsx':
        raise ValueError(f'Monster sheet must be .xlsx, got: {path.suffix}')
    types = typedefs_from_rows(iter_xlsx_rows(path))
    if register:
        for type_def in types:
            register_monster_type(type_def)
    return types


def load_default_monster_sheet():
    """Load project-root monster_types.xlsx. Returns [] if missing or unloadable."""
    if not DEFAULT_XLSX_PATH.is_file():
        print(
            f'[monster_types] missing {DEFAULT_XLSX_PATH}; '
            'only built-in species will spawn'
        )
        return []
    try:
        return load_monster_sheet(DEFAULT_XLSX_PATH)
    except ImportError as exc:
        print(
            f'[monster_types] cannot load {DEFAULT_XLSX_PATH}: {exc}. '
            'Install openpyxl in this Python environment '
            '(pip install openpyxl)'
        )
        return []
    except Exception as exc:
        print(f'[monster_types] failed to load {DEFAULT_XLSX_PATH}: {exc}')
        return []


def write_monster_xlsx(path=None, extra_rows=None):
    """Create the fill-in workbook (Monsters + Instructions + Field reference)."""
    from openpyxl import Workbook
    from openpyxl.comments import Comment
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    path = Path(path) if path else DEFAULT_XLSX_PATH
    wb = Workbook()

    thin = Border(
        left=Side(style='thin', color='C4B8A8'),
        right=Side(style='thin', color='C4B8A8'),
        top=Side(style='thin', color='C4B8A8'),
        bottom=Side(style='thin', color='C4B8A8'),
    )
    header_fill = PatternFill('solid', fgColor='3D2B1F')
    header_font = Font(bold=True, color='F4E8D0', name='Calibri', size=11)
    example_fill = PatternFill('solid', fgColor='E6F4EA')
    required_fill = PatternFill('solid', fgColor='3D2B1F')
    wrap = Alignment(wrap_text=True, vertical='center')

    ws = wb.active
    ws.title = 'Monsters'
    for col_idx, header in enumerate(COLUMNS, 1):
        cell = ws.cell(1, col_idx, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin
        note = HEADER_COMMENTS.get(header)
        if note:
            cell.comment = Comment(note, 'dungeon-console')

    data_rows = [TROLL_EXAMPLE]
    if extra_rows:
        data_rows.extend(extra_rows)
    blank_count = max(0, 40 - len(data_rows))

    for row_idx, row in enumerate(data_rows, 2):
        for col_idx, header in enumerate(COLUMNS, 1):
            cell = ws.cell(row_idx, col_idx, row.get(header, ''))
            cell.border = thin
            cell.alignment = wrap
            if row_idx == 2:
                cell.fill = example_fill

    last_data = 1 + len(data_rows) + blank_count
    for row_idx in range(2 + len(data_rows), last_data + 1):
        for col_idx in range(1, len(COLUMNS) + 1):
            cell = ws.cell(row_idx, col_idx, None)
            cell.border = thin

    widths = {
        'type_id': 14,
        'name': 16,
        'description': 42,
        'base_level': 12,
        'max_level': 12,
        'level_scaling': 14,
        'base_mhp': 12,
        'armour': 10,
        'ability_ids': 18,
        'sprite': 36,
        'portrait': 38,
        'spawn_notes': 36,
        'spawn_weight': 14,
        'sight_range': 13,
        'activeness': 12,
        'aggression': 12,
    }
    for col_idx, header in enumerate(COLUMNS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(header, 10)

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(COLUMNS))}{last_data}'
    ws.row_dimensions[1].height = 22
    ws.sheet_properties.tabColor = '3D2B1F'

    # Instructions
    ins = wb.create_sheet('Instructions')
    ins['A1'] = 'How to use'
    ins['A1'].font = Font(bold=True, size=14, color='3D2B1F')
    instructions = [
        '',
        '1. Fill one row per monster on the Monsters sheet (row 2 is the existing troll as an example).',
        '2. Required: type_id. Also fill name, description, the six attributes, base_mhp, and armour.',
        '3. armour: damage divisor (1 = full damage taken, 2 = half, etc.). Blank defaults to 1.',
        '4. Leave sprite/portrait blank to use default paths under static/monsters/.',
        '5. ability_ids: comma-separated list or blank. Abilities are data-only until combat hooks exist.',
        '6. spawn_weight: relative chance to appear. 0 means the species is registered but never random-spawned.',
        '7. max_level / level_scaling: dungeon spawns pick a random level 1..max_level; '
        f'bonus attribute points = (level - 1) * level_scaling (defaults {DEFAULT_MAX_LEVEL} / {DEFAULT_LEVEL_SCALING}).',
        '8. spawn_notes is for your own comments and is not loaded.',
        '9. Save this workbook, then restart the game. All rows with a type_id are imported together.',
        '',
        'AI field quick guide',
        'aggression 0-10: low flees, ~5 wanders, high chases.',
        'speed 0-10: 0 never gets a movement opportunity; any value above 0 acts once per world round.',
        'activeness 0-10: higher = less likely to stand still when idle or wandering.',
        'sight_range: Chebyshev vision distance in tiles.',
        '',
        'Art files',
        'Put PNG files at static/monsters/sprites/{type_id}.png and static/monsters/portraits/{type_id}.png,',
        'or set sprite/portrait to a custom path.',
    ]
    for idx, line in enumerate(instructions, 2):
        ins[f'A{idx}'] = line
        ins[f'A{idx}'].alignment = Alignment(wrap_text=True)
        if line in ('AI field quick guide', 'Art files'):
            ins[f'A{idx}'].font = Font(bold=True, size=12, color='3D2B1F')
    ins.column_dimensions['A'].width = 110

    # Field reference
    ref = wb.create_sheet('Field reference')
    ref_headers = ('column', 'required', 'type', 'notes')
    for col_idx, header in enumerate(ref_headers, 1):
        cell = ref.cell(1, col_idx, header)
        cell.fill = required_fill
        cell.font = header_font
        cell.border = thin
    fields = [
        ('type_id', 'yes', 'string', 'Unique machine id'),
        ('name', 'yes', 'string', 'Display name'),
        ('description', 'no', 'string', 'Inspect text'),
        ('base_level', 'no', 'int', 'Default 1 (used when level= omitted)'),
        ('max_level', 'no', 'int', f'Default {DEFAULT_MAX_LEVEL}. Random spawn upper bound'),
        ('level_scaling', 'no', 'int', f'Default {DEFAULT_LEVEL_SCALING}. Points per level above 1'),
        ('str/int/wis/chr/dex/agi', 'yes', 'int', 'Base attributes (missing keys become 1)'),
        ('base_mhp', 'yes', 'int', 'Max HP'),
        ('armour', 'no', 'int', 'Damage divisor. Default 1 (min 1)'),
        ('aggression', 'no', '0-10 float', 'Default 0'),
        ('speed', 'no', '0-10 float', 'Default 10. Only 0 vs >0 matters today'),
        ('activeness', 'no', '0-10 float', 'Default 5'),
        ('sight_range', 'no', 'int', 'Default 20'),
        ('ability_ids', 'no', 'csv string', 'Blank = none'),
        ('sprite', 'no', 'url path', 'Blank = /static/monsters/sprites/{type_id}.png'),
        ('portrait', 'no', 'url path', 'Blank = /static/monsters/portraits/{type_id}.png'),
        ('spawn_weight', 'no', 'float', 'Default 1. 0 = do not random-spawn'),
        ('spawn_notes', 'no', 'text', 'Ignored on import'),
    ]
    for row_idx, values in enumerate(fields, 2):
        for col_idx, value in enumerate(values, 1):
            cell = ref.cell(row_idx, col_idx, value)
            cell.border = thin
            cell.alignment = wrap
    ref.column_dimensions['A'].width = 28
    ref.column_dimensions['B'].width = 12
    ref.column_dimensions['C'].width = 16
    ref.column_dimensions['D'].width = 52
    ref.freeze_panes = 'A2'

    wb.save(path)
    return path
