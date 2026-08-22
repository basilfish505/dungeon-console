"""Load and write the armour types spreadsheet."""

from __future__ import annotations

from pathlib import Path

from armour_types.base import ArmourTypeDef
from armour_types.registry import register_armour_type

PROJECT_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_XLSX_PATH = PROJECT_ROOT / 'armour_types.xlsx'

COLUMNS = [
    'armour_id',
    'name',
    'description',
    'price_pqg',
    'image',
    'armour_value',
]

HEADER_COMMENTS = {
    'armour_id': 'Stable machine id, lowercase snake_case (e.g. leather). Required.',
    'name': 'Display name shown to players. Required.',
    'description': 'Inspect / inventory flavor text.',
    'price_pqg': 'Price in PermaQuest Gold (integer). Default 0.',
    'image': 'Sprite path or filename. Blank = /static/armour/sprites/{armour_id}.png',
    'armour_value': 'Damage divisor when equipped. Minimum 1. Default 1.',
}

TEST_ARMOUR = [
    {
        'armour_id': 'leather',
        'name': 'Leather Armour',
        'description': 'Light hide armour.',
        'price_pqg': 20,
        'image': 'leather.png',
        'armour_value': 2,
    },
    {
        'armour_id': 'chain_mail',
        'name': 'Chain Mail',
        'description': 'Interlocking metal rings.',
        'price_pqg': 60,
        'image': 'chain_mail.png',
        'armour_value': 3,
    },
    {
        'armour_id': 'plate',
        'name': 'Plate Armour',
        'description': 'Heavy steel plates.',
        'price_pqg': 120,
        'image': 'plate.png',
        'armour_value': 5,
    },
]


def _blank(value):
    if value is None:
        return True
    if isinstance(value, str) and not value.strip():
        return True
    return False


def _int(value, default):
    if _blank(value):
        return default
    return int(float(value))


def _str_or_none(value):
    if _blank(value):
        return None
    return str(value).strip()


def _normalize_header(value):
    if value is None:
        return ''
    return str(value).strip()


def row_to_typedef(row):
    armour_id = _str_or_none(row.get('armour_id'))
    if not armour_id:
        return None
    return ArmourTypeDef(
        armour_id=armour_id,
        name=_str_or_none(row.get('name')) or armour_id,
        description=_str_or_none(row.get('description')),
        price_pqg=_int(row.get('price_pqg'), 0),
        image=_str_or_none(row.get('image')),
        armour_value=_int(row.get('armour_value'), 1),
    )


def iter_xlsx_rows(path):
    from openpyxl import load_workbook

    wb = load_workbook(Path(path), data_only=True, read_only=True)
    try:
        ws = wb['Armour'] if 'Armour' in wb.sheetnames else wb.active
        rows = ws.iter_rows(values_only=True)
        header_row = next(rows, None)
        if not header_row:
            return
        headers = [_normalize_header(h) for h in header_row]
        for values in rows:
            yield {
                headers[i]: values[i] if i < len(values) else None
                for i in range(len(headers))
                if headers[i]
            }
    finally:
        wb.close()


def typedefs_from_rows(rows):
    types = []
    for row in rows:
        type_def = row_to_typedef(row)
        if type_def is not None:
            types.append(type_def)
    return types


def load_armour_sheet(path, register=True):
    path = Path(path)
    if not path.is_file():
        raise FileNotFoundError(path)
    if path.suffix.lower() != '.xlsx':
        raise ValueError(f'Armour sheet must be .xlsx, got: {path.suffix}')
    types = typedefs_from_rows(iter_xlsx_rows(path))
    if register:
        for type_def in types:
            register_armour_type(type_def)
    return types


def load_default_armour_sheet():
    if not DEFAULT_XLSX_PATH.is_file():
        print(
            f'[armour_types] missing {DEFAULT_XLSX_PATH}; '
            'no spreadsheet armour definitions loaded'
        )
        return []
    try:
        return load_armour_sheet(DEFAULT_XLSX_PATH)
    except ImportError as exc:
        print(
            f'[armour_types] cannot load {DEFAULT_XLSX_PATH}: {exc}. '
            'Install openpyxl (pip install openpyxl)'
        )
        return []
    except Exception as exc:
        print(f'[armour_types] failed to load {DEFAULT_XLSX_PATH}: {exc}')
        return []


def write_armour_xlsx(path=None, extra_rows=None):
    from openpyxl import Workbook
    from openpyxl.comments import Comment
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    path = Path(path) if path else DEFAULT_XLSX_PATH
    wb = Workbook()

    thin = Border(
        left=Side(style='thin', color='C4B8A8'),
        right=Side(style='thin', color='C4B8A8'),
        top=Side(style='thin', color='C4B8A8'),
        bottom=Side(style='thin', color='C4B8A8'),
    )
    header_fill = PatternFill('solid', fgColor='3D2B1F')
    header_font = Font(bold=True, color='F4E8D0', name='Calibri', size=11)
    example_fill = PatternFill('solid', fgColor='E6F4EA')
    wrap = Alignment(wrap_text=True, vertical='center')

    ws = wb.active
    ws.title = 'Armour'
    for col_idx, header in enumerate(COLUMNS, 1):
        cell = ws.cell(1, col_idx, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin
        note = HEADER_COMMENTS.get(header)
        if note:
            cell.comment = Comment(note, 'dungeon-console')

    data_rows = list(TEST_ARMOUR)
    if extra_rows:
        data_rows.extend(extra_rows)

    for row_idx, row in enumerate(data_rows, 2):
        for col_idx, header in enumerate(COLUMNS, 1):
            cell = ws.cell(row_idx, col_idx, row.get(header, ''))
            cell.border = thin
            cell.alignment = wrap
            if row_idx <= 4:
                cell.fill = example_fill

    blank_count = max(0, 40 - len(data_rows))
    last_data = 1 + len(data_rows) + blank_count
    for row_idx in range(2 + len(data_rows), last_data + 1):
        for col_idx in range(1, len(COLUMNS) + 1):
            cell = ws.cell(row_idx, col_idx, None)
            cell.border = thin

    widths = {
        'armour_id': 16,
        'name': 18,
        'description': 48,
        'price_pqg': 12,
        'image': 22,
        'armour_value': 14,
    }
    for col_idx, header in enumerate(COLUMNS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(header, 12)

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(COLUMNS))}{last_data}'
    ws.row_dimensions[1].height = 22

    ins = wb.create_sheet('Instructions')
    ins['A1'] = 'How to use'
    ins['A1'].font = Font(bold=True, size=14, color='3D2B1F')
    instructions = [
        '',
        '1. Fill one row per armour type on the Armour sheet.',
        '2. Required: armour_id. Also fill name, description, price_pqg, and armour_value.',
        '3. Leave image blank to use /static/armour/sprites/{armour_id}.png',
        '4. Save this workbook, then restart the game to reload definitions.',
        '',
        'Art files',
        'Put PNG files at static/armour/sprites/{armour_id}.png',
    ]
    for idx, line in enumerate(instructions, 2):
        ins[f'A{idx}'] = line
        if line in ('Art files',):
            ins[f'A{idx}'].font = Font(bold=True, size=12, color='3D2B1F')
    ins.column_dimensions['A'].width = 90

    wb.save(path)
    return path
