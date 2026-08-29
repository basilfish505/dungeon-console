"""Load spell types from the xlsx spreadsheet."""

import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from spell_types.base import SpellTypeDef, TARGET_MODES, _canon, _resolve_scaling_attribute
from spell_types.registry import SPELL_TYPES
from spell_types.sheet import (
    COLUMNS,
    load_spell_sheet,
    row_to_typedef,
    write_spell_xlsx,
)


class CanonTests(unittest.TestCase):
    def test_human_readable_target_and_hit(self):
        self.assertEqual(_canon('Single Enemy'), 'single_enemy')
        self.assertEqual(_canon('Always Hit'), 'always_hit')
        self.assertEqual(_canon('Damage'), 'damage')
        self.assertEqual(_canon('Single Any'), 'single_any')

    def test_scaling_attribute_from_label_or_key(self):
        self.assertEqual(_resolve_scaling_attribute('Intelligence'), 'int')
        self.assertEqual(_resolve_scaling_attribute('int'), 'int')
        self.assertEqual(_resolve_scaling_attribute('Strength'), 'str')

    def test_single_any_in_target_modes(self):
        self.assertIn('single_any', TARGET_MODES)


class ParseRowTests(unittest.TestCase):
    def test_skips_blank_spell_id(self):
        self.assertIsNone(row_to_typedef({'spell_id': '', 'name': 'Nope'}))
        self.assertIsNone(row_to_typedef({}))

    def test_parses_magic_bolt_row(self):
        td = row_to_typedef({
            'spell_id': 'magic_bolt',
            'name': 'Magic Bolt',
            'description': 'A bolt.',
            'effect_type': 'Damage',
            'target_mode': 'Single Enemy',
            'mp_cost': 2,
            'base_power': 5,
            'scaling_attribute': 'Intelligence',
            'scaling_factor': 1.0,
            'hit_rule': 'Always Hit',
            'spell_range': 6,
            'future_column': 'ignored',
        })
        self.assertEqual(td.id, 'magic_bolt')
        self.assertEqual(td.name, 'Magic Bolt')
        self.assertEqual(td.effect_type, 'damage')
        self.assertEqual(td.target_mode, 'single_enemy')
        self.assertEqual(td.mp_cost, 2)
        self.assertEqual(td.base_power, 5)
        self.assertEqual(td.scaling_attribute, 'int')
        self.assertEqual(td.scaling_factor, 1.0)
        self.assertEqual(td.hit_rule, 'always_hit')
        self.assertEqual(td.spell_range, 6)
        self.assertIsNone(td.min_power)
        self.assertIsNone(td.max_power)
        self.assertTrue(td.usable_in_combat)
        self.assertFalse(td.usable_out_of_combat)
        self.assertFalse(hasattr(td, 'future_column'))

    def test_parses_heal_row(self):
        td = row_to_typedef({
            'spell_id': 'heal',
            'name': 'Heal',
            'effect_type': 'heal',
            'target_mode': 'single_any',
            'mp_cost': 2,
            'min_power': 8,
            'max_power': 15,
            'hit_rule': 'always_hit',
            'usable_in_combat': 'yes',
            'usable_out_of_combat': 'yes',
        })
        self.assertEqual(td.id, 'heal')
        self.assertEqual(td.effect_type, 'heal')
        self.assertEqual(td.target_mode, 'single_any')
        self.assertEqual(td.min_power, 8)
        self.assertEqual(td.max_power, 15)
        self.assertTrue(td.usable_in_combat)
        self.assertTrue(td.usable_out_of_combat)

    def test_defaults_for_blank_cells(self):
        td = row_to_typedef({'spell_id': 'bare'})
        self.assertEqual(td.name, 'bare')
        self.assertEqual(td.effect_type, 'damage')
        self.assertEqual(td.target_mode, 'single_enemy')
        self.assertEqual(td.mp_cost, 0)
        self.assertEqual(td.base_power, 0)
        self.assertEqual(td.scaling_attribute, 'int')
        self.assertEqual(td.hit_rule, 'always_hit')
        self.assertTrue(td.usable_in_combat)
        self.assertFalse(td.usable_out_of_combat)


class XlsxLoadTests(unittest.TestCase):
    def test_xlsx_registers_magic_bolt_and_heal(self):
        previous = dict(SPELL_TYPES)
        try:
            with tempfile.TemporaryDirectory() as tmp:
                path = Path(tmp) / 'spells.xlsx'
                write_spell_xlsx(path)
                loaded = load_spell_sheet(path, register=True)
            ids = [td.id for td in loaded]
            self.assertEqual(ids, ['magic_bolt', 'heal'])
            bolt = SPELL_TYPES['magic_bolt']
            self.assertEqual(bolt.mp_cost, 2)
            self.assertEqual(bolt.base_power, 5)
            self.assertEqual(bolt.target_mode, 'single_enemy')
            self.assertFalse(bolt.usable_out_of_combat)
            heal = SPELL_TYPES['heal']
            self.assertEqual(heal.effect_type, 'heal')
            self.assertEqual(heal.target_mode, 'single_any')
            self.assertEqual(heal.min_power, 8)
            self.assertEqual(heal.max_power, 15)
            self.assertTrue(heal.usable_out_of_combat)
        finally:
            SPELL_TYPES.clear()
            SPELL_TYPES.update(previous)

    def test_xlsx_header_matches_schema(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / 'spells.xlsx'
            write_spell_xlsx(path)
            wb = load_workbook(path, read_only=True)
            try:
                ws = wb['Spells']
                header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            finally:
                wb.close()
        self.assertEqual(header, COLUMNS)

    def test_rejects_non_xlsx(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / 'spells.csv'
            path.write_text('spell_id,name\n', encoding='utf-8')
            with self.assertRaises(ValueError):
                load_spell_sheet(path, register=False)

    def test_typedef_construction_direct(self):
        td = SpellTypeDef(
            'test_bolt',
            name='Test',
            effect_type='damage',
            target_mode='single_enemy',
            mp_cost=2,
            base_power=5,
        )
        self.assertEqual(td.to_client_dict()['spell_id'], 'test_bolt')
        self.assertIn('usable_in_combat', td.to_client_dict())


if __name__ == '__main__':
    unittest.main()
