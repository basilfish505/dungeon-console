"""Load item types from the xlsx spreadsheet."""

import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from item_types.registry import ITEM_TYPES
from item_types.sheet import (
    COLUMNS,
    load_item_sheet,
    row_to_typedef,
    write_item_xlsx,
)


class ParseRowTests(unittest.TestCase):
    def test_skips_blank_item_id(self):
        self.assertIsNone(row_to_typedef({'item_id': '', 'name': 'Nope'}))
        self.assertIsNone(row_to_typedef({}))

    def test_parses_healing_potion_row(self):
        td = row_to_typedef({
            'item_id': 'healing_potion',
            'name': 'Healing Potion',
            'description': 'Restores a small amount of health.',
            'price_pqg': 25,
            'image': 'healing_potion.png',
            'future_column': 'ignored',
        })
        self.assertEqual(td.id, 'healing_potion')
        self.assertEqual(td.name, 'Healing Potion')
        self.assertEqual(td.description, 'Restores a small amount of health.')
        self.assertEqual(td.price_pqg, 25)
        self.assertTrue(td.image.endswith('/healing_potion.png'))

    def test_extra_column_ignored(self):
        td = row_to_typedef({
            'item_id': 'torch',
            'name': 'Torch',
            'category': 'tool',
            'weight': 1,
        })
        self.assertEqual(td.id, 'torch')
        self.assertFalse(hasattr(td, 'category'))


class XlsxLoadTests(unittest.TestCase):
    def test_xlsx_registers_test_items(self):
        previous = dict(ITEM_TYPES)
        try:
            with tempfile.TemporaryDirectory() as tmp:
                path = Path(tmp) / 'items.xlsx'
                write_item_xlsx(path)
                loaded = load_item_sheet(path, register=True)
            ids = [td.id for td in loaded]
            self.assertEqual(
                ids,
                [
                    'healing_potion',
                    'candle',
                    'torch',
                    'bread',
                    'rope',
                    'antidote',
                ],
            )
            self.assertEqual(ITEM_TYPES['healing_potion'].price_pqg, 25)
            self.assertEqual(ITEM_TYPES['torch'].name, 'Torch')
            self.assertEqual(ITEM_TYPES['candle'].light_sight, 1.5)
            self.assertEqual(ITEM_TYPES['torch'].light_ticks, 1000)
        finally:
            ITEM_TYPES.clear()
            ITEM_TYPES.update(previous)

    def test_xlsx_header_matches_schema(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / 'items.xlsx'
            write_item_xlsx(path)
            wb = load_workbook(path, read_only=True)
            try:
                ws = wb['Items']
                header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            finally:
                wb.close()
        self.assertEqual(header, COLUMNS)

    def test_rejects_non_xlsx(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / 'items.csv'
            path.write_text('item_id,name\n', encoding='utf-8')
            with self.assertRaises(ValueError):
                load_item_sheet(path, register=False)


if __name__ == '__main__':
    unittest.main()
