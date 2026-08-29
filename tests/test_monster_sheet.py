"""Load monster species from the xlsx spreadsheet."""

import random
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from monster_types.base import MonsterTypeDef
from monster_types.leveling import DEFAULT_LEVEL_SCALING, DEFAULT_MAX_LEVEL
from monster_types.registry import MONSTER_TYPES, pick_spawn_type_id, register_monster_type
from monster_types.sheet import (
    COLUMNS,
    DEFAULT_XLSX_PATH,
    load_monster_sheet,
    parse_ability_ids,
    row_to_typedef,
    write_monster_xlsx,
)


class ParseRowTests(unittest.TestCase):
    def test_skips_blank_type_id(self):
        self.assertIsNone(row_to_typedef({'type_id': '', 'name': 'Nope'}))
        self.assertIsNone(row_to_typedef({}))

    def test_parses_troll_like_row(self):
        td = row_to_typedef({
            'type_id': 'goblin',
            'name': 'Goblin',
            'description': 'Small and mean.',
            'base_level': 1,
            'str': 4,
            'int': 3,
            'wis': 2,
            'chr': 2,
            'dex': 6,
            'agi': 7,
            'base_mhp': 8,
            'aggression': 8,
            'speed': 10,
            'activeness': 7,
            'sight_range': 12,
            'ability_ids': 'stab, yell',
            'spawn_weight': 2,
        })
        self.assertEqual(td.id, 'goblin')
        self.assertEqual(td.name, 'Goblin')
        self.assertEqual(td.base_attributes['str'], 4)
        self.assertEqual(td.base_attributes['agi'], 7)
        self.assertEqual(td.base_mhp, 8)
        self.assertEqual(td.armour, 1)
        self.assertEqual(td.aggression, 8)
        self.assertEqual(td.activeness, 7)
        self.assertEqual(td.sight_range, 12)
        self.assertEqual(td.ability_ids, ['stab', 'yell'])
        self.assertEqual(td.spawn_weight, 2.0)
        self.assertTrue(td.sprite.endswith('/goblin.png'))
        self.assertEqual(td.level_scaling, DEFAULT_LEVEL_SCALING)
        self.assertEqual(td.max_level, DEFAULT_MAX_LEVEL)

    def test_parses_level_columns(self):
        td = row_to_typedef({
            'type_id': 'wisp',
            'name': 'Wisp',
            'max_level': 7,
            'level_scaling': 3,
            'str': 1,
            'int': 8,
            'wis': 8,
            'chr': 4,
            'dex': 5,
            'agi': 5,
            'base_mhp': 6,
        })
        self.assertEqual(td.max_level, 7)
        self.assertEqual(td.level_scaling, 3)

    def test_ability_ids_blank(self):
        self.assertEqual(parse_ability_ids(''), [])
        self.assertEqual(parse_ability_ids(None), [])
        self.assertEqual(parse_ability_ids('a, b ,c'), ['a', 'b', 'c'])

    def test_parses_spells_loot_and_base_mmp(self):
        td = row_to_typedef({
            'type_id': 'imp',
            'name': 'Imp',
            'str': 5,
            'int': 8,
            'wis': 8,
            'chr': 4,
            'dex': 6,
            'agi': 6,
            'base_mhp': 10,
            'spells': 'magic_bolt, frost_bolt',
            'loot': 'gold_coin, potion',
            'base_mmp': 6,
        })
        self.assertEqual(td.spell_ids, ['magic_bolt', 'frost_bolt'])
        self.assertEqual(td.loot_ids, ['gold_coin', 'potion'])
        self.assertEqual(td.base_mmp, 6)

    def test_spells_loot_default_empty(self):
        td = row_to_typedef({
            'type_id': 'rat',
            'name': 'Rat',
            'base_mhp': 4,
        })
        self.assertEqual(td.spell_ids, [])
        self.assertEqual(td.loot_ids, [])
        self.assertEqual(td.base_mmp, 0)


class XlsxLoadTests(unittest.TestCase):
    def test_xlsx_registers_multiple_types(self):
        extra = [{
            'type_id': 'goblin',
            'name': 'Goblin',
            'description': 'Small.',
            'base_level': 1,
            'str': 4,
            'int': 2,
            'wis': 2,
            'chr': 2,
            'dex': 6,
            'agi': 6,
            'base_mhp': 8,
            'armour': 2,
            'aggression': 8,
            'speed': 10,
            'activeness': 7,
            'sight_range': 12,
            'ability_ids': '',
            'sprite': '',
            'portrait': '',
            'spawn_weight': 2,
            'spawn_notes': '',
        }]
        previous = dict(MONSTER_TYPES)
        try:
            with tempfile.TemporaryDirectory() as tmp:
                path = Path(tmp) / 'monsters.xlsx'
                write_monster_xlsx(path, extra_rows=extra)
                loaded = load_monster_sheet(path, register=True)
            ids = [td.id for td in loaded]
            self.assertEqual(ids, ['troll', 'goblin'])
            self.assertEqual(MONSTER_TYPES['goblin'].base_mhp, 8)
            self.assertEqual(MONSTER_TYPES['goblin'].armour, 2)
            self.assertEqual(MONSTER_TYPES['goblin'].spawn_weight, 2.0)
            self.assertEqual(MONSTER_TYPES['troll'].name, 'Troll')
            self.assertEqual(MONSTER_TYPES['troll'].armour, 1)
        finally:
            MONSTER_TYPES.clear()
            MONSTER_TYPES.update(previous)

    def test_xlsx_header_matches_schema(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / 'monsters.xlsx'
            write_monster_xlsx(path)
            wb = load_workbook(path, read_only=True)
            try:
                ws = wb['Monsters']
                header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            finally:
                wb.close()
        self.assertEqual(header, COLUMNS)

    def test_rejects_non_xlsx(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / 'monsters.csv'
            path.write_text('type_id,name\n', encoding='utf-8')
            with self.assertRaises(ValueError):
                load_monster_sheet(path, register=False)

    def test_live_imp_row_has_magic_bolt_and_mp(self):
        if not DEFAULT_XLSX_PATH.is_file():
            self.skipTest('monster_types.xlsx missing')
        types = {td.id: td for td in load_monster_sheet(DEFAULT_XLSX_PATH, register=False)}
        self.assertIn('imp', types)
        imp = types['imp']
        self.assertEqual(imp.spell_ids, ['magic_bolt'])
        self.assertEqual(imp.base_mmp, 6)
        self.assertEqual(imp.loot_ids, [])


class SpawnPickTests(unittest.TestCase):
    def test_zero_weight_never_chosen(self):
        previous = dict(MONSTER_TYPES)
        try:
            MONSTER_TYPES.clear()
            register_monster_type(MonsterTypeDef(
                type_id='a', name='A', spawn_weight=0, base_mhp=1,
            ))
            register_monster_type(MonsterTypeDef(
                type_id='b', name='B', spawn_weight=5, base_mhp=1,
            ))
            rng = random.Random(0)
            picks = {pick_spawn_type_id(rng) for _ in range(20)}
            self.assertEqual(picks, {'b'})
        finally:
            MONSTER_TYPES.clear()
            MONSTER_TYPES.update(previous)

    def test_empty_registry_falls_back_to_troll(self):
        previous = dict(MONSTER_TYPES)
        try:
            MONSTER_TYPES.clear()
            self.assertEqual(pick_spawn_type_id(), 'troll')
        finally:
            MONSTER_TYPES.update(previous)


if __name__ == '__main__':
    unittest.main()
