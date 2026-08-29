"""Load and write the spell types spreadsheet."""

from __future__ import annotations

from pathlib import Path

from spell_types.base import SpellTypeDef, _optional_int, _parse_bool
from spell_types.registry import register_spell_type

PROJECT_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_XLSX_PATH = PROJECT_ROOT / 'spell_types.xlsx'

COLUMNS = [
    'spell_id',
    'name',
    'description',
    'effect_type',
    'target_mode',
    'mp_cost',
    'base_power',
    'scaling_attribute',
    'scaling_factor',
    'min_power',
    'max_power',
    'hit_rule',
    'spell_range',
    'usable_in_combat',
    'usable_out_of_combat',
]

HEADER_COMMENTS = {
    'spell_id': 'Stable machine id, lowercase snake_case (e.g. magic_bolt). Required.',
    'name': 'Display name shown to players. Required.',
    'description': 'Inspect / picker flavor text.',
    'effect_type': (
        'damage | heal | buff | debuff | status | utility. '
        'damage and heal are implemented; others load but cannot cast yet.'
    ),
    'target_mode': (
        'single_enemy | single_any | self | single_ally | all_enemies | all_allies. '
        'single_enemy and single_any are implemented.'
    ),
    'mp_cost': 'Mana points spent when the spell is successfully cast. Default 0.',
    'base_power': 'Flat power added before attribute scaling (damage). Default 0.',
    'scaling_attribute': (
        'Attribute that scales power: int / Intelligence, or any ATTRIBUTE_KEYS label. '
        'Default int.'
    ),
    'scaling_factor': 'Multiplier applied to the scaling attribute. Default 1.0.',
    'min_power': (
        'Inclusive low end of a flat roll (e.g. heal). Blank = use base_power scaling.'
    ),
    'max_power': (
        'Inclusive high end of a flat roll. Blank = use base_power scaling.'
    ),
    'hit_rule': 'always_hit | accuracy. Only always_hit is used so far. Default always_hit.',
    'spell_range': 'Max Chebyshev distance in tiles (informational for now). Default 1.',
    'usable_in_combat': 'yes/no. Default yes.',
    'usable_out_of_combat': 'yes/no. Default no.',
}

TEST_SPELLS = [
    {
        'spell_id': 'magic_bolt',
        'name': 'Magic Bolt',
        'description': 'A simple bolt of force. Test spell for the magic system.',
        'effect_type': 'damage',
        'target_mode': 'single_enemy',
        'mp_cost': 2,
        'base_power': 5,
        'scaling_attribute': 'int',
        'scaling_factor': 1.0,
        'min_power': '',
        'max_power': '',
        'hit_rule': 'always_hit',
        'spell_range': 6,
        'usable_in_combat': 'yes',
        'usable_out_of_combat': 'no',
    },
    {
        'spell_id': 'heal',
        'name': 'Heal',
        'description': 'Restore hit points to any living target.',
        'effect_type': 'heal',
        'target_mode': 'single_any',
        'mp_cost': 2,
        'base_power': 0,
        'scaling_attribute': 'int',
        'scaling_factor': 0,
        'min_power': 8,
        'max_power': 15,
        'hit_rule': 'always_hit',
        'spell_range': 6,
        'usable_in_combat': 'yes',
        'usable_out_of_combat': 'yes',
    },
]


def _blank(value):
    if value is None:
        return True
    if isinstance(value, str) and not value.strip():
        return True
    return False


def _int(value, default):
    if _blank(value):
        return default
    return int(float(value))


def _float(value, default):
    if _blank(value):
        return default
    return float(value)


def _str_or_none(value):
    if _blank(value):
        return None
    return str(value).strip()


def _normalize_header(value):
    if value is None:
        return ''
    return str(value).strip()


def row_to_typedef(row):
    spell_id = _str_or_none(row.get('spell_id'))
    if not spell_id:
        return None
    return SpellTypeDef(
        spell_id=spell_id,
        name=_str_or_none(row.get('name')) or spell_id,
        description=_str_or_none(row.get('description')),
        effect_type=_str_or_none(row.get('effect_type')) or 'damage',
        target_mode=_str_or_none(row.get('target_mode')) or 'single_enemy',
        mp_cost=_int(row.get('mp_cost'), 0),
        base_power=_int(row.get('base_power'), 0),
        scaling_attribute=_str_or_none(row.get('scaling_attribute')) or 'int',
        scaling_factor=_float(row.get('scaling_factor'), 1.0),
        hit_rule=_str_or_none(row.get('hit_rule')) or 'always_hit',
        spell_range=_int(row.get('spell_range'), 1),
        min_power=_optional_int(row.get('min_power')),
        max_power=_optional_int(row.get('max_power')),
        usable_in_combat=_parse_bool(row.get('usable_in_combat'), True),
        usable_out_of_combat=_parse_bool(row.get('usable_out_of_combat'), False),
    )


def iter_xlsx_rows(path):
    from openpyxl import load_workbook

    wb = load_workbook(Path(path), data_only=True, read_only=True)
    try:
        ws = wb['Spells'] if 'Spells' in wb.sheetnames else wb.active
        rows = ws.iter_rows(values_only=True)
        header_row = next(rows, None)
        if not header_row:
            return
        headers = [_normalize_header(h) for h in header_row]
        for values in rows:
            yield {
                headers[i]: values[i] if i < len(values) else None
                for i in range(len(headers))
                if headers[i]
            }
    finally:
        wb.close()


def typedefs_from_rows(rows):
    types = []
    for row in rows:
        type_def = row_to_typedef(row)
        if type_def is not None:
            types.append(type_def)
    return types


def load_spell_sheet(path, register=True):
    path = Path(path)
    if not path.is_file():
        raise FileNotFoundError(path)
    if path.suffix.lower() != '.xlsx':
        raise ValueError(f'Spell sheet must be .xlsx, got: {path.suffix}')
    types = typedefs_from_rows(iter_xlsx_rows(path))
    if register:
        for type_def in types:
            register_spell_type(type_def)
    return types


def load_default_spell_sheet():
    if not DEFAULT_XLSX_PATH.is_file():
        print(
            f'[spell_types] missing {DEFAULT_XLSX_PATH}; '
            'no spreadsheet spell definitions loaded'
        )
        return []
    try:
        return load_spell_sheet(DEFAULT_XLSX_PATH)
    except ImportError as exc:
        print(
            f'[spell_types] cannot load {DEFAULT_XLSX_PATH}: {exc}. '
            'Install openpyxl (pip install openpyxl)'
        )
        return []
    except Exception as exc:
        print(f'[spell_types] failed to load {DEFAULT_XLSX_PATH}: {exc}')
        return []


def write_spell_xlsx(path=None, extra_rows=None):
    from openpyxl import Workbook
    from openpyxl.comments import Comment
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    path = Path(path) if path else DEFAULT_XLSX_PATH
    wb = Workbook()

    thin = Border(
        left=Side(style='thin', color='C4B8A8'),
        right=Side(style='thin', color='C4B8A8'),
        top=Side(style='thin', color='C4B8A8'),
        bottom=Side(style='thin', color='C4B8A8'),
    )
    header_fill = PatternFill('solid', fgColor='3D2B1F')
    header_font = Font(bold=True, color='F4E8D0', name='Calibri', size=11)
    example_fill = PatternFill('solid', fgColor='E6F4EA')
    wrap = Alignment(wrap_text=True, vertical='center')

    ws = wb.active
    ws.title = 'Spells'
    for col_idx, header in enumerate(COLUMNS, 1):
        cell = ws.cell(1, col_idx, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin
        note = HEADER_COMMENTS.get(header)
        if note:
            cell.comment = Comment(note, 'dungeon-console')

    data_rows = list(TEST_SPELLS)
    if extra_rows:
        data_rows.extend(extra_rows)

    for row_idx, row in enumerate(data_rows, 2):
        for col_idx, header in enumerate(COLUMNS, 1):
            cell = ws.cell(row_idx, col_idx, row.get(header, ''))
            cell.border = thin
            cell.alignment = wrap
            if row_idx <= 4:
                cell.fill = example_fill

    blank_count = max(0, 40 - len(data_rows))
    last_data = 1 + len(data_rows) + blank_count
    for row_idx in range(2 + len(data_rows), last_data + 1):
        for col_idx in range(1, len(COLUMNS) + 1):
            cell = ws.cell(row_idx, col_idx, None)
            cell.border = thin

    widths = {
        'spell_id': 16,
        'name': 16,
        'description': 48,
        'effect_type': 12,
        'target_mode': 14,
        'mp_cost': 10,
        'base_power': 12,
        'scaling_attribute': 18,
        'scaling_factor': 14,
        'min_power': 12,
        'max_power': 12,
        'hit_rule': 12,
        'spell_range': 12,
        'usable_in_combat': 16,
        'usable_out_of_combat': 18,
    }
    for col_idx, header in enumerate(COLUMNS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(header, 12)

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(COLUMNS))}{last_data}'
    ws.row_dimensions[1].height = 22

    ins = wb.create_sheet('Instructions')
    ins['A1'] = 'How to use'
    ins['A1'].font = Font(bold=True, size=14, color='3D2B1F')
    instructions = [
        '',
        '1. Fill one row per spell on the Spells sheet.',
        '2. Required: spell_id. Also fill name, effect_type, target_mode, mp_cost.',
        '3. Damage spells use base_power + scaling. Heal uses min_power/max_power roll.',
        '4. usable_in_combat / usable_out_of_combat: yes or no.',
        '5. Save this workbook, then restart the game to reload definitions.',
        '',
        'Implemented now',
        'damage + single_enemy (Magic Bolt); heal + single_any (Heal).',
    ]
    for idx, line in enumerate(instructions, 2):
        ins[f'A{idx}'] = line
        if line in ('Implemented now',):
            ins[f'A{idx}'].font = Font(bold=True, size=12, color='3D2B1F')
    ins.column_dimensions['A'].width = 90

    wb.save(path)
    return path
